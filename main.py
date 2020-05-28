# -*- coding: utf-8 -*-
import config
import number_pi

import telebot
import openpyxl

bot = telebot.TeleBot(config.token)

class Student:
    def __init__(self):
        self.name = None
        self.surname = None
        self.group = 0
        self.number_in_list = None

student = Student()

command_start = config.command_start
com = config.com

@bot.message_handler(commands=['reg', 'help'])
def registration(message):
    if message.text == "/reg":
        global student
        student = Student()
        global isRunning
        if not isRunning:
            msg = bot.send_message(message.from_user.id, "Введи свое фамилию и имя через пробел.\nПример: Иванов Иван")
            bot.register_next_step_handler(msg, get_name)
            isRunning = True
    elif message.text == "/help":
        bot.send_message(message.from_user.id, f"Доступные комманды:\n{com}")

@bot.message_handler(commands=['info'])
def info(message):
    if student.group != 0:
        bot.send_message(message.chat.id, student.name + " " + student.surname + " " + str(student.group))
    else:
        bot.send_message(message.chat.id, "Ты ещё не зарегистрирован!\nНапиши /reg")

@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, message.from_user.first_name + command_start)

group = 0

@bot.message_handler(commands=['my_tasks'])
def my_tasks(message):
    global group
    if student.group == 0:
        bot.send_message(message.from_user.id, 'Для начала необходимо зарегистрироваться!\nНапиши /reg')
        group = 0
        return
    msg = bot.send_message(message.from_user.id, 'Введи первый номер и последний')
    bot.register_next_step_handler(msg, get_problems)

@bot.message_handler(commands=['183_task'])
def tasks_183(message):
    global group
    group = 183
    my_tasks(message)

@bot.message_handler(commands=['185_task'])
def tasks_185(message):
    global group
    group = 185
    my_tasks(message)

@bot.message_handler(commands=['186_task'])
def tasks_186(message):
    global group
    group = 186
    my_tasks(message)

@bot.message_handler(commands=['187_task'])
def tasks_187(message):
    global group
    group = 187
    my_tasks(message)

@bot.message_handler(commands=['188_task'])
def tasks_188(message):
    global group
    group = 188
    my_tasks(message)

def get_problems(message):
    global group
    if not group:
        group = student.group

    start_problem, end_problem = [int(x) for x in message.text.strip().split()]
    nums = 'Твои номера:'
    numbers = []
    for n_task in range(start_problem, end_problem + 1):
        N = (n_task - 1) * 300 + (student.group - 183) * 35 + student.number_in_list
        nums += " " + str(n_task) + "." + str(number_pi.pi[N - 1])
        numbers.append(number_pi.pi[N - 1])
    bot.send_message(message.from_user.id, nums)
    j = 0
    answer = ""
    for n_task in range(start_problem, end_problem + 1):
        database = openpyxl.load_workbook(filename="./data.xlsx")
        sheet = database[str(group)]
        answer += f"Задание {n_task}.{numbers[j]} совпадает сo студентами:\n\n"
        i = 2
        res = ""
        while sheet['A' + str(i)].value != None:
            N_stud = (n_task - 1) * 300 + (group - 183) * 35 + int(sheet['A' + str(i)].value)
            current_name = sheet['B' + str(i)].value.strip()
            if numbers[j] == number_pi.pi[N_stud - 1] and int(sheet['A' + str(i)].value) != student.number_in_list:
                if not res:
                    res = current_name
                else:
                    res += ", " + current_name
            i += 1
        if not res:
            answer += "Нет совпадений\n\n"
        else:
            answer += res + "\n\n"
        j += 1
    bot.send_message(message.from_user.id, answer)
    group = 0

def find_group(message):
    bot.send_message(message.chat.id, "Выполняется поиск студента")
    database = openpyxl.load_workbook(filename="./data.xlsx")
    global student
    for gr in database.sheetnames:
        sheet = database[gr]
        i = 2
        while sheet['B' + str(i)].value != None and student.group == 0:
            current_name = sheet['B' + str(i)].value.strip().split()
            if current_name[0] == student.surname and current_name[1] == student.name:
                student.group = int(gr)
                student.number_in_list = int(sheet['A' + str(i)].value)
            i += 1
        if student.group:
            break
    if student.group != 0:
        bot.send_message(message.chat.id, f"Тебя зовут {student.name} {student.surname}."
        f"Твоя группа {student.group}. Твой номер в списке {student.number_in_list}")
    else:
        bot.send_message(message.chat.id, f"Проверь свои данные. Ты ввел\nИмя: {student.name}. Фамилия: {student.surname}")

isRunning = False

@bot.message_handler(content_types=['text'])
def get_text_messages(message):
    bot.send_message(message.from_user.id, "Я тебя не понял. Посмотри доступные команды с помощью /help")


def get_name(message):
    global student
    name = message.text
    if name.isdigit():
        msg = bot.send_message(message.from_user.id, 'Фамилия и Имя не могут быть числом! Введи ещё раз!')
        bot.register_next_step_handler(msg, get_name)
        return
    student.surname, student.name = [x for x in message.text.strip().split()]
    global isRunning
    isRunning = False
    find_group(message)

def get_surname(message):
    surname = message.text
    if surname.isdigit():
        msg = bot.send_message(message.from_user.id, 'Фамилия не может быть числом! Введите ещё раз!')
        bot.register_next_step_handler(msg, get_surname)
        return
    student.surname = surname
    global isRunning
    isRunning = False
    find_group(message)

bot.polling(none_stop=True, interval=0)